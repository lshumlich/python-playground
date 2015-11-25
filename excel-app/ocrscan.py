from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import utils
import re

"""
LeaseType    LeaseID    RightsGranted    RoyaltyScheme    CrownMultiplier    MinRoyalty    ValuationMethod    TruckingDeducted    ProcessingDeducted    Notes
OL    1    All    SKProvCrownVar    1.2    50    SaskWellHead    Y    Y    
"""

class ParseLease(object):
    
    def __init__(self,getter):
        self.getCellValue = getter
    
    def leaseType(self):
        lease = self.getCellValue('Leaseno')
        return lease [0:2]
        
    def leaseID(self):
        lease = self.getCellValue('Leaseno')
        return int (lease [3:9])
    
    def leaseRights(self):
        leaseRights = self.getCellValue('Lsrights')
        rightsGranted = "*** Research Required ***"
        if leaseRights == "Oil & Gas [Excluding Crude Bitumen]":
            rightsGranted = "O+G-CBit"
        if leaseRights == "GasOnlyOilOnly":
            rightsGranted = "O+G-CBit"
        if leaseRights == "All Rights [O&G including Bitumen]":
            rightsGranted = "All+Bit"
        if leaseRights == "Oil & Gas [Excluding Crude Bitumen] + Disposal":
            rightsGranted = "O+G-CBit-Disp"    
        return rightsGranted
    
    def oilMultiplier(self):
        oilMultiplierP = re.compile('oil royalty rate means ([0-9.]+) times', re.I)
        appBDetail = self.getCellValue('Royappb')
        oilMultiplier = None
        if (appBDetail != None):
            oilMultiplier = '*** Research Required ***'
            m = oilMultiplierP.search(appBDetail)
            if m != None:
                oilMultiplier = m.group(1)
        return oilMultiplier

    def royaltyScheme(self):
        regulationP = re.compile('regulation royalty rate',re.I)
        crownVar1P = re.compile('multiples of provincial',re.I)
        crownVar2P = re.compile('modified provincial crown',re.I)
        gorrP = re.compile('gorr',re.I)
        s = self.getCellValue('Royschem')
        royaltyScheme = ''
        if s != None:
            royaltyScheme = self.lookForString(s,regulationP,royaltyScheme,'IOGR1995')
            royaltyScheme = self.lookForString(s,crownVar1P,royaltyScheme,'SKProvCrownVar')
            royaltyScheme = self.lookForString(s,crownVar2P,royaltyScheme,'SKProvCrownVar')
            royaltyScheme = self.lookForString(s,gorrP,royaltyScheme,'GORR')
        if royaltyScheme == '':
            royaltyScheme += '*** Research Required ***'
        return royaltyScheme

    def gorr(self):
        percentP = re.compile('[0-9.]+%')
        s = self.getCellValue('Gorr')
        gorrStuff = None
        if s != None:
            gorrStuff = percentP.findall(s)
        return gorrStuff
        
    def lookForString (self,sToLookIn,pattern,result,sToAppend):
            m = pattern.search(sToLookIn)
            if m != None:
                if result != '':
                    result +=','
                result += sToAppend
            return result

class MultiLineWS(object):

    def getCellValue(self,label):
        
        i = 0
        for cell in self.headerRow:
            if cell.value == label:
                return self.row[i].value
            i += 1
        raise Exception('Cell ' + label + ' not found.')
             
    def process(self,wsName):
        
        parseLease = ParseLease(self.getCellValue)
        
        leasewb = load_workbook(wsName)
        ws = leasewb.active
        
        self.headerRow = None
        i = 0
         
        for self.row in ws.rows:
            i += 1
            if self.headerRow == None:
                self.headerRow = self.row
            else:
                print('row:', i, 'LeaseType:',parseLease.leaseType(),'leaseID:',parseLease.leaseID(),
                      'leaseRights:',parseLease.leaseRights(),
                      'oilMultiplier:',parseLease.oilMultiplier(),
                      'royaltyScheme:',parseLease.royaltyScheme(),
                      'gorr:',parseLease.gorr())
    

class SingeLeaseWS(object):

    def getCellValue(self,label):
        
        for row in self.ws.rows:
            if row[0].value == label:
                return row[1].value
        raise Exception('Cell ' + label + ' not found.')
             
    def process(self,wsName):
        
        parseLease = ParseLease(self.getCellValue)
        
        leasewb = load_workbook(wsName)
        self.ws = leasewb.active

        print('LeaseType:',parseLease.leaseType())
        print('leaseID:',parseLease.leaseID())
        print('leaseRights:',parseLease.leaseRights())
        print('oilMultiplier:',parseLease.oilMultiplier())
        print('royaltyScheme:',parseLease.royaltyScheme())
        print('gorr:',parseLease.gorr())
#
# Sample of running with the Multi Line WS        
#
# multiLine = MultiLineWS()
# multiLine.process(r"D:\code\info\Lease Analysis - Sask Only.xlsx")
#
# Sample of running with Single Lease WS
#
singleLease = SingeLeaseWS()
singleLease.process(r"OL-0001.xlsx")
