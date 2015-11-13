#!/bin/env python3

"""
The next step is to write a generic worksheet loader that will load any
type of data from a tab. The top row must be the field names, the rest of
the rows are the data. 

"""
import unittest
from openpyxl import load_workbook
from openpyxl import Workbook
from apperror import AppError

class DataStructure(object):
    
    def __init__(self):
        None
        
    def __str__(self):
        return str(vars(self))
    
    def headers(self):
        d = vars(self)
        return list(d.keys())

    def data(self):
        d = vars(self)
        return list(d.values())

"""
    Generic class for fetching data.
    This is to simulate a database but much much much easier to change
"""

class DataBase(object):
    
    def __init__(self,worksheetName):
        try:
            self.worksheetName = worksheetName
            self.wb = load_workbook(worksheetName)
            self.wellTabName = 'Well'
            self.royaltyMasterTabName = 'RoyaltyMaster'
            self.leaseTabName = 'Lease'
            self.monthlyTabName = 'Monthly'
            self.producingEntityTabName = 'ProducingEntity'
            self.monthlyRecordId = 0
            self.productClausesTabName = 'ProductClauses'
            self.econOilDataTabName = 'ECONData'
        except FileNotFoundError:
            raise AppError('The excel worksheet ' + worksheetName + ' is not found')
        self.loadWellFromExcel()
        self.loadRoyaltyMasterFromExcel()
        self.loadLeaseFromExcel()
        self.loadProductClausesFromExcel()
        self.loadMonthlyFromExcel()
        self.loadProducingEntityFromExcel()
        self.loadECONOilDataFromExcel()
        
    #
    # Royalty Master
    #
    def loadRoyaltyMasterFromExcel(self):
        stack = self.excelLoadWsTable(self.royaltyMasterTabName)
        self.royaltyMaster = dict()
        for ds in stack:
            self.royaltyMaster[ds.Lease] = ds
    
    def getRoyaltyMaster(self, lease):
        try:
            return self.royaltyMaster[lease]
        except KeyError:
            raise AppError ('Royalty Master not found for Lease: ' + lease)
        
    #
    # Lease
    #
    def loadLeaseFromExcel(self):
        stack = self.excelLoadWsTable(self.leaseTabName)
        self.lease = dict()
        for ds in stack:
            self.lease[ds.Lease] = ds
        
    def getLease(self, lease):
        try:
            return self.lease[lease]
        except KeyError:
            raise AppError ('Lease not found for Lease: ' + lease)
    #
    # Well
    #
    def loadWellFromExcel(self):
        stack = self.excelLoadWsTable(self.wellTabName)
        self.well = dict()
        for ds in stack:
            self.well[ds.WellId] = ds
        
    def getWell(self, wellId):
        try:
            return self.well[wellId]
        except KeyError:
            raise AppError ('Well not found for WellId: ' + str(wellId))
    #
    # Producing Entity 
    #
    def loadProducingEntityFromExcel(self):
        stack = self.excelLoadWsTable(self.producingEntityTabName)
        self.producingEntity = dict()
        for ds in stack:
            self.producingEntity[ds.PEID] = ds
        
    def getProducingEntity(self, lease):
        try:
            return self.producingEntity[lease]
        except KeyError:
            raise AppError ('Producing Entity not found for Lease: ' + lease)
    #
    # Product Clauses
    #
    def loadProductClausesFromExcel(self):
        stack = self.excelLoadWsTable(self.productClausesTabName)
        self.productClauses = dict()
        for ds in stack:
            self.royaltyMaster[ds.Lease] = ds

    #
    # Monthly Data 
    #
    def loadMonthlyFromExcel(self):
        self.monthlyTable = self.excelLoadWsTable(self.monthlyTabName)
        self.monthly = iter(self.monthlyTable)
    
    def monthlyDataNextRow(self):
        return next(self.monthly)

    def monthlyData(self):
        return self.monthly
    
    def getMonthlyData(self):
        return self.monthlyTable[0]

    #
    # ECON Monthly Oil Factors Data
    #
    def loadECONOilDataFromExcel(self):
        stack = self.excelLoadWsTable(self.econOilDataTabName)
        self.econOilData = dict()
        for ds in stack:
            self.econOilData[ds.ProdMonth] = ds
        
    def getECONOilData(self, prodMonth):
        try:
            return self.econOilData[prodMonth]
        except KeyError:
            raise AppError ('ECONOilData not found for: ' + str(prodMonth))

    #
    # Royalty Calculation
    #
    def getRoyaltyCalc(self,month,wellId):
        rc = DataStructure()
        setattr(rc, 'ProdMonth', month)
        setattr(rc, 'WellId', wellId)
        setattr(rc, 'K', 0.0)
        setattr(rc, 'X', 0.0)
        setattr(rc, 'C', 0.0)
        setattr(rc, 'D', 0.0)
        setattr(rc, 'RoyaltyRate', 0.0)
        setattr(rc, 'CalcRoyaltyRate', 0.0)
        setattr(rc, 'RoyaltyPrice', 0.0)
        setattr(rc, 'RoyaltyVolume', 0.0)
        setattr(rc, 'RoyaltyValuePreDeductions', 0.0)
        setattr(rc, 'RoyaltyTransportation', 0.0)
        setattr(rc, 'RoyaltyProcessing', 0.0)
        setattr(rc, 'RoyaltyDeductions', 0.0)
        setattr(rc, 'RoyaltyValue', 0.0)
        setattr(rc, 'CommencementPeriod', None)
        

        
        return rc

    def updateRoyaltyCalc(self, rc):
        None
    #
    # Generic load a tab into a data structure
    #
    def excelLoadWsTable(self,tabName):
        try: 
            ws = self.wb[tabName]
            stack = []
            recordNo = 0;
            headerRow = None
            for row in ws.rows:
                if headerRow == None:
                    headerRow = row
                else:
                    recordNo = recordNo +1
                    ds = DataStructure()
                    stack.append(ds)
                    i = 0
                    for cell in headerRow:
                        setattr(ds, cell.value, row[i].value)
                        i = i + 1
                    setattr(ds, 'RecordNumber', recordNo)
        except KeyError:
            raise AppError('The excel worksheet ' + self.worksheetName + ' does not have tab: ' + tabName)
        except TypeError as e:
            print('Error Loading tab:',tabName,' column:',i,'Record:',recordNo,'Error:',e)
            print('   headerRow:',headerRow)
            print('         row:',row)
            raise e
            
        return stack
    
    #
    # Used for testing
    #
    def forceWorkBook(self, workBook):
        self.wb = workBook
##    def royaltyMaster(self, id):
##        try:
##            return self.royaltyMasterData[id]
##        except AttributeError as e:
##            print('*** Time to load the royalty master buddy***')
            
"""
    Test Code....

    For these test to work there must be an excel file named "sample.xlsx"
    with the correct tabs.
"""
class TestDataBase(unittest.TestCase):
    
    def setUp(self):
        self.validExcelFile = 'database.xlsx'
        
    def test_getWorkSheetThatDoesNotExist(self):
        self.assertRaises(AppError,DataBase,'no way this exists.xlsxx')
        
    def test_excelLoadWsTable(self):
        fd = DataBase(self.validExcelFile)
        fd.excelLoadWsTable()
        
    def test_excelLoadWsTableTabFales(self):
        fd = DataBase(self.validExcelFile)
        fd.forceWorkBook(Workbook()) # force an emply workbook
        self.assertRaises(AppError,fd.excelLoadWsTable,'SomeTabName')
        
    def test_readingMonthlyData(self):
        DataBase(self.validExcelFile)
        
    def test_getRoyaltyMasterNotFound(self):
        fd = DataBase(self.validExcelFile)
        self.assertRaises(AppError,fd.getRoyaltyMaster,'BadLease')

    def test_somethingelse(self):
        DataBase(self.validExcelFile)

    def getDataBase(self):
        """ for test purposes get a database instance """
        return DataBase(self.validExcelFile)


#print ('Looking for ID 3 - ', self.royaltyMaster[3])
#print ('Looking for ID 500 - ', self.royaltyMaster[500])


"""
The crux of each test is a call to assertEqual() to check for an expected
result; assertTrue() or assertFalse() to verify a condition; or assertRaises()
to verify that a specific exception gets raised. These methods are used instead
of the assert statement so the test runner can accumulate all test results and
produce a report.

"""

if __name__ == '__main__':
    unittest.main()

##d = {1:0, 2:0, 3:0}
##print(d.keys())
##print(list(d.keys()))
##print(list(d.values()))
