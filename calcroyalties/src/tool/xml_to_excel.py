#!/bin/env python3
"""
This is responsible for converting xml data files into to excel. We then will
load the excel into sql tables.

As of today this is just a hack file

"""

from openpyxl import Workbook
import xml.etree.ElementTree as et
import os
import config

# Global Variables

wb = Workbook()
ws = None
order = 0
row = 2

# ws = wb.create_sheet("MySheet")

# print ('---Tag-->', root.tag)
# print('----attr->', root.attrib)
#
# for c in root.iter():
#     print(c.tag, " - ", c.attrib, c.text, ' - ' )


def output_structure(e, level, name):
    global order, row
    tab_order = {}

    if name:
        name += '.' + e.tag
    else:
        name = e.tag

    dup_children = False
    if len(e) > 1:
        if e[0].tag == e[1].tag:
            dup_children = True

    if len(e) == 0:
        order += 1
        update_cell(name, row, order, e.text)

    # print(' '*level*3, level, row, order, "me--> ", name, ' ', e.attrib, ' ', e.text, ' Length=',len(e), dup_children)

    if len(e) > 0:
        for i in range(0, len(e)):
            if e[i].tag in tab_order:
                order = tab_order[e[i].tag]
                row += 1
            else:
                if i < len(e) - 1:
                    if e[i].tag == e[i+1].tag:
                        tab_order[e[i].tag] = order

            # print('-->', order)
            output_structure(e[i], level+1, name)

def write_excel():
    print("About to write some excel.")
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("Mysheet")
    ws1.cell(row=1, column=1, value=12345)
    wb.save(config.get_temp_dir() + 'lfsdata.xlsx')


def create_sheet(name):
    global ws, wb
    ws = wb.create_sheet(name)


def update_cell(attr, row, col, value):
    global ws
    ws.cell(row=row, column=col, value=value)
    if not ws.cell(row=1, column=col).value:
        ws.cell(row=1, column=col, value=attr)

def process_xml_file(name):
    global order, row
    print("process_xml_file --> ", name)


    tree = et.parse(name)
    root = tree.getroot()
    e = root        # Envelope
    e = e[0]        # Body
    e = e[0]        # LIST_FACILITY_003
    e = e[1]        # DATAAREA

    create_sheet(e[0].tag)
    # e = e[0]    # First Child (In this case facility)
    order = 0
    row = 2
    output_structure(e, 0, None)

def process_xml_dir(name):
    for f in os.listdir(name):
        if f.endswith('.xml'):
            # print(name+f)
            process_xml_file(name+f)

# process_xml_file('\\$temp\\xml\\Sample-VOLUMETRIC.xml')
# tree = et.parse('\\$temp\\xml\\Sample-FACILITY.xml')
# process_xml_file(config.get_temp_dir() + 'Sample-FACILITY.xml')
process_xml_dir(config.get_temp_dir())


# print('Length=',len(e),' Length=',len(e[0]))

# e = get_dataarea()

# print(e)
# show_children(e)
#
# e = e.firstChild
# show_children(e)
#


excel_file = config.get_temp_dir() + 'lfsdata.xlsx'
print("Writing Excel file:", excel_file)
wb.save(excel_file)



#
#
# cd = CommonData()
#
# # showStruc(root, 0)
# ns = {'soap':'http://schemas.xmlsoap.org/soap/envelope/'}
# e = root
# e = e.find('soap:Body', ns)
# e = e.find('LIST_FACILITY_003')
# e = e.find('DATAAREA')
# print('length:',len(e))
# print('data-->',e[0].tag, e[1].tag)
# print('data-->',e[0][0].tag, e[1][0].tag)
# show_children(e)

