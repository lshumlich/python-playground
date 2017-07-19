import openpyxl
import config


def update_id():

    wb = openpyxl.load_workbook(config.get_temp_dir() + 'sample_data.xlsx')
    d = {}

    well_royalty_master = wb.get_sheet_by_name('WellRoyaltyMaster')
    monthly = wb.get_sheet_by_name('Monthly')
    monthly['R1'] = 'WellEvent'
    entity_lease_link = wb.get_sheet_by_name('EntityLeaseLink')
    entity_lease_link['H1'] = 'WellEvent'

    for row in range(2, well_royalty_master.max_row + 1):
        id = well_royalty_master["A" + str(row)].value
        well_event = well_royalty_master["D" + str(row)].value
        d[id] = well_event

    for row in range(2, monthly.max_row + 1):
        well_id = monthly["E" + str(row)].value
        if well_id in d.keys():
            monthly['R' + str(row)] = d[well_id]

    for row in range(2, entity_lease_link.max_row +1):
        well_id = entity_lease_link["E" + str(row)].value
        if well_id in d.keys():
            entity_lease_link['H' + str(row)] = d[well_id]

    wb.save('sample_data_new.xslx')


update_id()